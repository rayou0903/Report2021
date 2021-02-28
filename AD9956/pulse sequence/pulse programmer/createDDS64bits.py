# -*- coding: utf-8 -*-

import openpyxl as px

#fname = '/Users/yokooannosuke/Cording/Pine64-python--pulse/Pulseprogramer/ver2/pulsedata.xlsm'
fname = 'C:/Users/rayou/Python/pulsedata.xlsm'

wb = px.load_workbook(fname)
# print(type(wb))  # Bookオブジェクトを取得
sheet = wb['AD9956パラメータ']  # 変更

##########################################################################################
# 入力された周波数（f）を48bitdataに変換

col_values0 = []
Fout = []
settingF = []
Fout_new = []
f_idx = 0
f_sys = sheet.cell(column=1,row=4).value  # REFCLK周波数
div_r = sheet.cell(column=2,row=4).value  # 分周比

for n in range(10):
    if n % 5 == 0: # n=0,5のみ実行
        for cell_obj in list(sheet.columns)[n]:
            col_values0.append(cell_obj.value)
        del col_values0[0:6]  # 関係ないセル(値以外)を削除　セルの変更に伴い変更
        del col_values0[8]
        Fout = [s for s in col_values0 if s != '']  # Fout内の’’を削除
        Fout_new.append(Fout)
        col_values0 = []

while f_idx < 2:
    for frq in Fout_new[f_idx]:
        # 設定周波数 ＝（2^48/クロック周波数）＊ 出力周波数
        setF = (pow(2, 48) // f_sys) * frq * div_r  # 式変更
        setF = int(setF)  # 小数点以下切り捨て
        set48bitF = format(setF, 'b')  # 2進数に変換 変数名変更(以降すべてに適用)
        set48bitF = str(set48bitF).zfill(48)  # 右寄せ0埋め ビット数変更
        # print(set48bitF)
        settingF.append(set48bitF)  # 要素の追加，データの一時保存
        setF, set48bitF = 0, ' '  # 初期化
    f_idx += 1

frq_data1, frq_data2 = settingF[0:8], settingF[8: 16]  # DDS1,DDS2用に分割する
frq_data = [frq_data1, frq_data2]
# print(frq_data)

##############################################################################
# 位相制御5bit

Pout = []
Pout_new = []
settingP = []
col_values1 = []
p_idx = 0

for n in range(10):
    if n % 5 == 1: # n=1,6のみ実行
        for cell_obj in list(sheet.columns)[n]:
            col_values1.append(cell_obj.value)
        del col_values1[0:6]  # セルの変更に伴い変更
        del col_values1[8]
        Pout = [s for s in col_values1 if s != '']
        Pout_new.append(Pout)
        col_values1 = []

while p_idx < 2:
    for phase in Pout_new[p_idx]:
        θ = phase
        setP = (θ * pow(2, 14)) / 360  # 式変更
        setP = int(setP)
        set14bitP = format(setP, 'b')  # 変数名変更(以降すべてに適用)
        set14bitP = str(set14bitP)
        set14bitP = set14bitP.zfill(14)
        # print(set14bitP)
        settingP.append(set14bitP)
        setP, set14bitP = 0, ' '
    p_idx += 1

pha_data1, pha_data2 = settingP[0:8], settingP[8:16]  # DDS1,DDS2用に分割する
pha_data = [pha_data1, pha_data2]
# print(pha_data)

######################################################################################
# パワーダウン　ロジック　6＊REFの3bit部分
"""
REFout = []
REFout_new = []
settingREF = []
col_values2 = []
REF_idx = 0

for n in range(12):
    if n % 6 == 2:
        for cell_obj in list(sheet.columns)[n]:
            col_values2.append(cell_obj.value)
        del col_values2[0:3]
        REFout = [s for s in col_values2 if s != '']
        REFout_new.append(REFout)
        col_values2 = []

while REF_idx < 2:
    for ref in REFout_new[REF_idx]:
        a = str(ref)
        settingREF.append(a.zfill(3))
    REF_idx += 1

power_data1, power_data2 = settingREF[0:8], settingREF[8:16]
power_data = [power_data1, power_data2]
# print(power_data)
"""
#####################################################################################

# 各部分の統合
# Phase5bit　REF3bit Frequency32bit　の順番

DDS64bitdata0 = []  # 変数名変更
DDS64bitdata1 = []  # DDS1の方の40bitdata
DDS64bitdata2 = []  # DDS2の方の40bitdata
index = 0

while index < 2:
    for n in range(len(frq_data[index])):
        bitdata64 = '00' + pha_data[index][n] + frq_data[index][n]  # AD9956用に変更
        DDS64bitdata0.append(bitdata64)
    index += 1
    # print(DDS40bitdata0)


DDS64bitdata1 = DDS64bitdata0[0:8]  # スライスで分割
DDS64bitdata2 = DDS64bitdata0[8: 16]

print(DDS64bitdata1[0])
print(DDS64bitdata1[7])
print(DDS64bitdata2[0])
print(DDS64bitdata2[7])


#########################################################################################
# excelファイルに書き込み
# いったんファイルを閉じて開かないと更新されない

wb = px.load_workbook(fname, keep_vba=True)
sheet = wb['AD9956パラメータ']  # 変更

# 書き込むセルの位置を変更(colums→ABC..., row→123...)
for m in range(len(DDS64bitdata1)):
    sheet.cell(column=3, row=(m + 7), value=DDS64bitdata1[m])

for n in range(len(DDS64bitdata2)):
    sheet.cell(column=8, row=(n + 7), value=DDS64bitdata2[n])

wb.save(fname)
print('finished!')
print('何かキーを押してください')
input()